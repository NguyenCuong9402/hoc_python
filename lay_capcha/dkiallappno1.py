import requests
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
import time
import requests
import re
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
import os
import subprocess
import json
import random
import urllib.request
import socket
import urllib.error
from openpyxl import Workbook,load_workbook
import urllib.request
import pytesseract
from PIL import Image
import base64
from PIL import Image, ImageEnhance, ImageFilter
from io import BytesIO
import cv2
import string
import requests



def gettxt(file,dong):
 with open(file) as f:
    contents = f.readlines()
    sl=len(contents)
    return contents[dong]
def settxt(file,noidung):
 f = open(file, "a")
 f.write('\n'+noidung)
app=gettxt('app.txt',0)
def getexecl(file,cot):
    wb =load_workbook(file)
    ws =wb.active
    x=cot
    return ws[x].value
def updatefile(file,cot,noidung):
    wb =load_workbook(file)
    ws =wb.active
    x=cot
    ws[x].value=noidung
    wb.save(file)
    return True
def is_bad_proxy(pip):    
    try:
        proxy_handler = urllib.request.ProxyHandler({'http': pip})
        opener = urllib.request.build_opener(proxy_handler)
        opener.addheaders = [('User-agent', 'Mozilla/5.0')]
        urllib.request.install_opener(opener)
        req=urllib.request.Request('http://www.example.com')  # change the URL to test here
        sock=urllib.request.urlopen(req)
    except urllib.error.HTTPError as e:
        return e.code
    except Exception as detail:
        return True
    return False
def fakeip(key):
  # a=requests.get('http://proxy.tinsoftsv.com/api/changeProxy.php?key='+key+'&location='+id1+'')
  PROXY=''
  while True:
    try:
      data={'key': key}
      a=requests.get('https://app.proxyno1.com/api/change-key-ip/'+key,timeout=15)
      data = json.loads(a.text)
      kq=data['status']
      if int(kq)==0:
        break
      else :
        time.sleep(0.1)
    except Exception as ex:
       print('Lỗi',ex)
       time.sleep(0.1)
def download_file(url, dst_path):
    try:
        with urllib.request.urlopen(url) as web_file:
            data = web_file.read()
            with open(dst_path, mode='wb') as local_file:
                local_file.write(data)
    except urllib.error.URLError as e:
        print(e)

def demo_imagetotext(Base64):
    url='data:image/png;base64,'+Base64
    i=random.randint(1, 999)
    dst_path=current_dir+'/image'+str(i)+'.png'
    download_file(url, dst_path)
    img = Image.open(dst_path)
    text = pytesseract.image_to_string(img)
    os.remove(dst_path)
    text=text.replace('/','').replace("\n",'')
    return text[:4]
def random_char(y):
       return ''.join(random.choice(string.ascii_letters) for x in range(y))
def merge_name_to_lowercase(name):
    # Xóa bỏ dấu cách trong tên
    name_without_spaces = name.replace(" ", "")
    # Chuyển thành chữ thường
    lowercase_name = name_without_spaces.lower()
    a=len(lowercase_name)
    so = random.randint(6, 9)
    sosau = random.randint(10, 99)
    if a<8:
            chu=random_char(3)
            ten=lowercase_name[:a]+chu.lower()+str(sosau)
    else:
            chu=random_char(2)
            ten=lowercase_name[:so]+chu.lower()+str(sosau)

    return ten

def checkip(session,proxies):
    ip=session.get('https://ipinfo.io/json',proxies=proxies, timeout=5)
    ip=json.loads(ip.text)
    ipcuatoi=ip['ip']
    return ipcuatoi
def header(so,fake3):
    headers ={
    'accept': 'application/json, text/plain, */*',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'en-US,en;q=0.9,vi;q=0.8',
    'content-language': 'vi-VN',
    'content-length': '433',
    'content-type': 'application/json;charset=UTF-8',
    'cookie': 'NG_TRANSLATE_LANG_KEY=vi; GuestVersion=2156716; MemberVersion=2156716; MarqueeVersion=53; NeverPopupDatetime=2023-04-10%2009%3A37%3A21; nohostname_ip=3734AB59AG12664276531E; AWSALB=vgksxRt9FlN6YIOIbTrtlkP1M2x0+/LPsvtE0V9WyVDalz/rv8DZ/ER5dSCfux2SAQ10N3m/t3eLO3l7O5AVBHtj5rlEYyMrM6WpJr+vYk5jeRENcv6qKKHbUcf8; AWSALBCORS=vgksxRt9FlN6YIOIbTrtlkP1M2x0+/LPsvtE0V9WyVDalz/rv8DZ/ER5dSCfux2SAQ10N3m/t3eLO3l7O5AVBHtj5rlEYyMrM6WpJr+vYk5jeRENcv6qKKHbUcf8',
    'origin': 'https://m.99win55.com',
    'referer': 'https://m.99win55.com/',
    'sec-ch-ua': '"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': "Mozilla/5.0 (iPhone; CPU iPhone OS 15_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/104.0.5112."+str(so)+" Mobile/21E"+str(so)+" Safari/"+str(fake3)+".1",
    'x-requested-with': 'XMLHttpRequest'
    }
    return headers
def random_char(y):
       return ''.join(random.choice(string.ascii_letters) for x in range(y))

# data1={'image': "/9j/4AAQSkZJRgABAQEAYABgAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAASACgDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD0u4u002c3GoXtvbW5IXzJ5AiHrgDJ69aa2oQ2xF/c6haxWUgBW4eVVjcHlcEnFUb9Zf7TmvrQWdxf2UZV/OuSgijYZ5UA8nAOCB9fXndKik0ddO1uO1nu9Ii07dG7mNGgLEszFXYAZBx8pOMY5B589JNf15nFGOsdJfE3v/Wn/B11Oun1D7Gv255fOhkC+V5Pz+fu+6qAdSf/AK5OOafay3waW61BI7SNQXbfKCFj6/Meikc5wSB6muKtdIuY00KfWNOmlsFNxLLbRwvMImkJZFMajcQAWXocE81PDbzabpzxz6PdPbXV81xa2ccUuyCMMGTeI1YoMjdtwcFsY603Far+v+HCKsou0vib39dfT/g6nX6deR6jcPPZ31vdQA7WkgkDr9ODwef855KzfDlvM+o3+pu8hmunXzt9pJbqAq4VVVwGOB/F3/Sis6mjO7BK0JWvu9/0OG8ZXNxF8U7OwjnlSzu5LYXNurkRzZbB3r0bIGOe1da4E/xUSymAktYdME0UD8okgkADqvQMASARzRRWsdomT+L5v9TqI2b+05Rk42DjP0ojZv7TlGTjYOM/SiisSo/Z/wAb/UI2b+05Rk42DjP0ooopM3w/wy9X+Z//2Q==",
# 'value': "06KfGRXCOLM5FhCBv5cwcA=="}


# URL của API
url = "https://www.8kwin.cc/api/0.0/Home/GetCaptchaForLogin"

# Gửi yêu cầu POST
response = requests.post(url)

# Kiểm tra xem yêu cầu có thành công hay không
if response.status_code == 200:
    data = response.json()
    print("Image:", data.get("image"))
    print("Value:", data.get("value"))
    data1 = {
        "image": data.get("image"), "value": data.get("value")}
else:
    print("Failed to retrieve captcha. Status code:", response.status_code)




url = 'https://'+app+'/'
print('Web Dang ky ',url)
urlcaptcha='https://'+app+'/api/0.0/Home/GetCaptchaForRegister'

current_dir = os.path.dirname(os.path.abspath(__file__))
mk=gettxt('mk.txt',0)
file=input('Nhập file execl:').replace('"','')
key=input('Nhập key no1:')
ipcuatoi=input('Nhap proxy:')
print('ip của bạn là:',ipcuatoi)
cot=int(input('Cột:'))
j=0
filelichsu=file.replace('.xlsx','').replace(' ','')+str(cot)+'kq.txt'
f = open(filelichsu, 'w')
f.close()
while True:
 try:
    fakeip(key)
    proxies=ipcuatoi
    session= requests.Session()
    proxies = {'https': proxies}
    while True:
        ip=checkip(session,proxies)
        so = random.randint(1, 99)
        fake3=random.randint(100, 999)
        headers=header(so,fake3)
        apicaptcha=session.post(urlcaptcha,headers=headers,json=data1,proxies=proxies, timeout=5)
        base64 = json.loads(apicaptcha.text)['image']
        checkCodeEncrypt = json.loads(apicaptcha.text)['value']
        codecaptcha=demo_imagetotext(base64)


 except Exception as ex:
    print('[',cot,']=>',ex)
    
    continue



